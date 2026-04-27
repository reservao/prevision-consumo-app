from flask import Flask, request, send_file, render_template_string
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io, os, tempfile, xml.etree.ElementTree as ET
from datetime import datetime

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

HTML = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Previsión de Consumo</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Serif+Display&family=DM+Sans:wght@300;400;500;600&display=swap" rel="stylesheet">
<style>
  *{box-sizing:border-box;margin:0;padding:0}
  :root{--navy:#1a2e4a;--blue:#1F4E79;--accent:#2e86de;--light:#EBF5FB;--mid:#D6E4F0;--muted:#6b7c93;--white:#fff;--success:#27ae60;--error:#e74c3c}
  body{font-family:'DM Sans',sans-serif;background:#f0f4f8;min-height:100vh;display:flex;align-items:center;justify-content:center;padding:2rem}
  .card{background:var(--white);border-radius:24px;box-shadow:0 20px 60px rgba(26,46,74,.12);padding:3rem 3.5rem;max-width:560px;width:100%}
  .logo-row{display:flex;align-items:center;gap:.75rem;margin-bottom:2rem}
  .logo-icon{width:44px;height:44px;background:linear-gradient(135deg,var(--blue),var(--accent));border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:1.4rem}
  .logo-text{font-family:'DM Serif Display',serif;font-size:1.1rem;color:var(--navy);line-height:1.2}
  .logo-text span{display:block;font-family:'DM Sans',sans-serif;font-size:.75rem;color:var(--muted);letter-spacing:.04em;text-transform:uppercase}
  h1{font-family:'DM Serif Display',serif;font-size:2rem;color:var(--navy);line-height:1.2;margin-bottom:.75rem}
  .subtitle{color:var(--muted);font-size:.95rem;line-height:1.6;margin-bottom:2.5rem}
  .steps{display:flex;gap:1rem;margin-bottom:2.5rem}
  .step{flex:1;background:var(--light);border-radius:12px;padding:1rem;text-align:center}
  .step-num{font-size:1.4rem;margin-bottom:.3rem}
  .step-label{font-size:.75rem;color:var(--blue);font-weight:600}
  .drop-zone{border:2px dashed var(--mid);border-radius:16px;padding:2.5rem 2rem;text-align:center;cursor:pointer;transition:all .25s;background:var(--light);margin-bottom:1.5rem;position:relative}
  .drop-zone:hover,.drop-zone.dragover{border-color:var(--accent);background:rgba(46,134,222,.05)}
  .drop-zone input[type=file]{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%}
  .drop-icon{font-size:2.5rem;margin-bottom:.75rem}
  .drop-title{font-weight:600;color:var(--navy);font-size:.95rem;margin-bottom:.3rem}
  .drop-sub{font-size:.8rem;color:var(--muted)}
  .file-selected{background:rgba(39,174,96,.08);border-color:var(--success);border-style:solid}
  .file-name{font-size:.85rem;color:var(--success);font-weight:600;margin-top:.5rem;word-break:break-all}
  .btn{width:100%;padding:1rem;background:linear-gradient(135deg,var(--blue),var(--accent));color:#fff;border:none;border-radius:12px;font-family:'DM Sans',sans-serif;font-size:1rem;font-weight:600;cursor:pointer;transition:all .2s;box-shadow:0 4px 16px rgba(31,78,121,.25)}
  .btn:hover:not(:disabled){transform:translateY(-2px)}
  .btn:disabled{opacity:.6;cursor:not-allowed}
  .alert{padding:1rem 1.25rem;border-radius:10px;font-size:.88rem;margin-top:1rem;display:none}
  .alert.error{background:rgba(231,76,60,.1);color:var(--error);border:1px solid rgba(231,76,60,.2);display:block}
  .alert.success{background:rgba(39,174,96,.1);color:var(--success);border:1px solid rgba(39,174,96,.2);display:block}
  .loading{display:none;align-items:center;justify-content:center;gap:.75rem;padding:1rem;color:var(--blue);font-size:.9rem;font-weight:500}
  .loading.show{display:flex}
  .spinner{width:20px;height:20px;border:3px solid var(--mid);border-top-color:var(--accent);border-radius:50%;animation:spin .8s linear infinite}
  @keyframes spin{to{transform:rotate(360deg)}}
  .footer{text-align:center;margin-top:2rem;font-size:.75rem;color:var(--muted)}
</style>
</head>
<body>
<div class="card">
  <div class="logo-row">
    <div class="logo-icon">📊</div>
    <div class="logo-text">Previsión de Consumo<span>Transformador de datos</span></div>
  </div>
  <h1>Transforma tu sábana de datos</h1>
  <p class="subtitle">Sube tu archivo .xls con los datos apilados y descarga automáticamente una tabla Excel ordenada y categorizada.</p>
  <div class="steps">
    <div class="step"><div class="step-num">📂</div><div class="step-label">Sube el .xls</div></div>
    <div class="step"><div class="step-num">⚡</div><div class="step-label">Procesamos</div></div>
    <div class="step"><div class="step-num">⬇️</div><div class="step-label">Descarga</div></div>
  </div>
  <form id="uploadForm">
    <div class="drop-zone" id="dropZone">
      <input type="file" id="fileInput" name="file" accept=".xls,.xlsx">
      <div class="drop-icon">📄</div>
      <div class="drop-title">Arrastra tu archivo aquí</div>
      <div class="drop-sub">o haz clic para seleccionar — .xls / .xlsx</div>
      <div class="file-name" id="fileName"></div>
    </div>
    <button class="btn" type="submit" id="submitBtn" disabled>Transformar y descargar Excel</button>
  </form>
  <div class="loading" id="loading"><div class="spinner"></div>Procesando tu archivo, un momento…</div>
  <div class="alert" id="alertBox"></div>
  <div class="footer">Los archivos se procesan en memoria y no se almacenan en ningún servidor.</div>
</div>
<script>
  const fi=document.getElementById('fileInput'),fn=document.getElementById('fileName'),
        sb=document.getElementById('submitBtn'),dz=document.getElementById('dropZone'),
        ab=document.getElementById('alertBox'),ld=document.getElementById('loading'),
        fm=document.getElementById('uploadForm');
  fi.addEventListener('change',()=>{if(fi.files.length>0){fn.textContent='✅ '+fi.files[0].name;dz.classList.add('file-selected');sb.disabled=false;}});
  dz.addEventListener('dragover',e=>{e.preventDefault();dz.classList.add('dragover');});
  dz.addEventListener('dragleave',()=>dz.classList.remove('dragover'));
  dz.addEventListener('drop',e=>{e.preventDefault();dz.classList.remove('dragover');const f=e.dataTransfer.files;if(f.length>0){fi.files=f;fn.textContent='✅ '+f[0].name;dz.classList.add('file-selected');sb.disabled=false;}});
  fm.addEventListener('submit',async(e)=>{
    e.preventDefault();ab.className='alert';ab.textContent='';ld.classList.add('show');sb.disabled=true;
    const fd=new FormData();fd.append('file',fi.files[0]);
    try{
      const r=await fetch('/transformar',{method:'POST',body:fd});
      if(r.ok){const b=await r.blob(),u=URL.createObjectURL(b),a=document.createElement('a');a.href=u;a.download='Prevision_Consumo_Tabla.xlsx';a.click();URL.revokeObjectURL(u);ab.className='alert success';ab.textContent='✅ ¡Archivo transformado y descargado exitosamente!';}
      else{const err=await r.json();ab.className='alert error';ab.textContent='❌ Error: '+(err.error||'Ocurrió un problema.');}
    }catch(err){ab.className='alert error';ab.textContent='❌ Error de conexión. Intenta de nuevo.';}
    finally{ld.classList.remove('show');sb.disabled=false;}
  });
</script>
</body></html>"""


def leer_xml_spreadsheetml(filepath):
    NS = 'urn:schemas-microsoft-com:office:spreadsheet'
    tree = ET.parse(filepath)
    root = tree.getroot()
    rows_data = []
    for worksheet in root.findall(f'{{{NS}}}Worksheet'):
        table = worksheet.find(f'{{{NS}}}Table')
        if table is None:
            continue
        for row in table.findall(f'{{{NS}}}Row'):
            cells = []
            col_index = 0
            for cell in row.findall(f'{{{NS}}}Cell'):
                idx_attr = cell.get(f'{{{NS}}}Index')
                if idx_attr is not None:
                    target = int(idx_attr) - 1
                    while col_index < target:
                        cells.append(None)
                        col_index += 1
                data = cell.find(f'{{{NS}}}Data')
                if data is not None and data.text:
                    val = data.text.strip()
                    dtype = data.get(f'{{{NS}}}Type', 'String')
                    if dtype == 'Number':
                        try:
                            val = float(val)
                        except ValueError:
                            pass
                    cells.append(val)
                else:
                    cells.append(None)
                col_index += 1
            rows_data.append(cells)
    if not rows_data:
        raise ValueError("No se encontraron filas en el XML")
    max_cols = max(len(r) for r in rows_data)
    for r in rows_data:
        while len(r) < max_cols:
            r.append(None)
    return pd.DataFrame(rows_data), True


def transformar_archivo(filepath):
    df = None
    use_xml = False

    try:
        df = pd.read_excel(filepath, header=None, engine='openpyxl')
    except Exception:
        pass
    if df is None:
        try:
            df = pd.read_excel(filepath, header=None, engine='xlrd')
        except Exception:
            pass
    if df is None:
        try:
            df, use_xml = leer_xml_spreadsheetml(filepath)
        except Exception:
            pass
    if df is None:
        raise ValueError("No se pudo leer el archivo. Asegúrate de subir un .xls o .xlsx válido.")

    col_cant = 1 if use_xml else 3
    col_uni  = 2 if use_xml else 8

    records = []
    unit_markers = df[df[0].isin(['Unidad Agregada', 'Unidade Agregada'])].index.tolist()
    for idx, marker in enumerate(unit_markers):
        unit_id   = str(df.iloc[marker + 1, 0]).strip()
        unit_name = str(df.iloc[marker + 2, 0]).strip()
        end = unit_markers[idx + 1] if idx + 1 < len(unit_markers) else len(df)
        i = marker + 3
        while i < end:
            codigo   = str(df.iloc[i, 0]).strip()
            cantidad = df.iloc[i, col_cant] if df.shape[1] > col_cant else None
            unidad   = str(df.iloc[i, col_uni]).strip() if df.shape[1] > col_uni else ''
            nombre   = str(df.iloc[i + 1, 0]).strip() if i + 1 < end else ''
            if '.' in codigo and pd.notna(cantidad):
                try:
                    records.append({
                        'ID Unidad Agregada':     unit_id,
                        'Nombre Unidad Agregada': unit_name,
                        'Código Producto':        codigo,
                        'Nombre Producto':        nombre,
                        'Cantidad Bruta':         float(cantidad),
                        'Unidad Medida':          unidad
                    })
                except (ValueError, TypeError):
                    pass
            i += 2

    if not records:
        raise ValueError("No se encontraron datos válidos. Verifica que el archivo tenga el formato correcto.")

    result = pd.DataFrame(records)

    # Generar Excel con pandas (rápido) + formato mínimo solo en encabezados
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        result.to_excel(writer, index=False, sheet_name='Previsión de Consumo')
        ws = writer.sheets['Previsión de Consumo']

        # Formato encabezados (solo 6 celdas — muy rápido)
        for col_idx in range(1, 7):
            cell = ws.cell(row=1, column=col_idx)
            cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            cell.fill      = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Anchos de columna
        for i, w in enumerate([18, 35, 18, 38, 16, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Formato numérico cantidad
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                cell.number_format = '#,##0.000'

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:F{len(result) + 1}'

    output.seek(0)
    return output


@app.route('/')
def index():
    return render_template_string(HTML)


@app.route('/transformar', methods=['POST'])
def transformar():
    if 'file' not in request.files:
        return {'error': 'No se recibió ningún archivo.'}, 400
    file = request.files['file']
    if file.filename == '':
        return {'error': 'El archivo está vacío.'}, 400
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ['.xls', '.xlsx']:
        return {'error': 'Solo se aceptan archivos .xls o .xlsx'}, 400
    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        file.save(tmp.name)
        tmp_path = tmp.name
    try:
        output = transformar_archivo(tmp_path)
        return send_file(output,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name='Prevision_Consumo_Tabla.xlsx')
    except Exception as e:
        return {'error': str(e)}, 500
    finally:
        os.unlink(tmp_path)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
